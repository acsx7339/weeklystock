import openpyxl
from datetime import datetime, timedelta
import twstock


class ReviewExcel:
    def __init__(self):
        self.file_name = "stock_review.xlsx"
        self.wb = openpyxl.load_workbook(self.file_name)

        # 計算過去 7 天的日期，並格式化為工作表名稱
        seven_days_ago = datetime.now() - timedelta(days=7)
        self.past_date = seven_days_ago.strftime("%Y-%m-%d")
        self.sheet_name = self.past_date

        # 確認是否存在對應的工作表
        if self.sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' does not exist in the workbook!")

    def get_price(self):
        sheet = self.wb[self.sheet_name]

        # 找到 "stock" 欄位的索引
        header_row = 1  # 假設標題在第一行
        name_column_index = None
        for col_index, cell in enumerate(sheet[header_row], start=1):
            if cell.value == "stock":
                name_column_index = col_index
                break

        if name_column_index is None:
            raise ValueError("Column with header 'stock' not found!")

        # 提取 "stock" 欄位的值
        stock_items = [
            sheet.cell(row=row, column=name_column_index).value
            for row in range(2, sheet.max_row + 1)  # 從第 2 行開始提取
        ]

        stock_prices = []
        for item in stock_items:
            try:
                stock = twstock.Stock(item)
                past_price = stock.price[-10]  # 提取 10 天前的價格
                stock_prices.append(past_price)
            except Exception as e:
                stock_prices.append(None)  # 如果有錯誤，記錄為 None
                print(f"Error fetching price for stock {item}: {e}")

        return stock_prices

    def update_prices(self, prices):
        sheet = self.wb[self.sheet_name]

        # 找到 "Close" 欄位所在列，或創建一個新的列
        header = "Close"
        header_column = sheet.max_column + 1  # 預設在最後一列後新增
        for col_index, cell in enumerate(sheet[1], start=1):  # 檢查第一行
            if cell.value == header:
                header_column = col_index
                break

        # 添加標題 "Close"（如果是新列）
        sheet.cell(row=1, column=header_column, value=header)

        # 更新價格資料
        for row, value in enumerate(prices, start=2):  # 從第 2 行開始
            sheet.cell(row=row, column=header_column, value=value)

        # 儲存工作簿
        self.wb.save(self.file_name)

        print(f"Updated prices in column '{header_column}' with values: {prices}")


# if __name__ == "__main__":
#     try:
#         rw = ReviewExcel()
#         prices = rw.get_price()
#         rw.update_prices(prices)
#     except Exception as e:
#         print(f"An error occurred: {e}")
