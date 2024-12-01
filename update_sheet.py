import openpyxl
from datetime import datetime, timedelta
import twstock


class UpdateExcel:
    def __init__(self):
        self.file_name = "stock_review.xlsx"
        self.wb = openpyxl.load_workbook(self.file_name)
        self.current_date = datetime.now().strftime("%Y-%m-%d")  # 當天日期
        self.past_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")  # 七天前日期

        # 如果當天的工作表不存在，創建新工作表
        if self.current_date in self.wb.sheetnames:
            self.sheet = self.wb[self.current_date]
            print(f"使用現有的工作表：{self.current_date}")
        else:
            self.sheet = self.wb.create_sheet(title=self.current_date)
            print(f"已創建新的工作表：{self.current_date}")

    def add_item(self, sdict):
        """
        將股票及價格資料新增到 Excel。
        """
        if not sdict:
            print("沒有資料可添加！")
            return

        # 添加標題行（如果是空工作表）
        if self.sheet.max_row == 1:
            self.sheet.append(["stock", "StartPrice"])

        # 添加股票與價格資料
        for stock, price in sdict.items():
            self.sheet.append([stock, price])

        # 儲存工作簿
        self.wb.save(self.file_name)
        print(f"資料已成功新增到工作表：{self.current_date}")

    def get_price(self, stock_items):
        """
        獲取股票清單的最新價格。
        """
        stock_dict = {}
        for item in stock_items:
            try:
                stock = twstock.Stock(item)
                current_price = stock.price[-1]
                stock_dict[item] = current_price
            except Exception as e:
                print(f"無法獲取股票 {item} 的價格: {e}")
                stock_dict[item] = None  # 標記為無法獲取價格
        return stock_dict


# if __name__ == "__main__":
    # 測試流程
    # e = UpdateExcel()

    # 測試用股票清單
    # stock_list = ['2059', '2330', '2357', '2414', '2545', '3037', '3454', '6024', '6515', '6807']
    # stock_prices = e.get_price(stock_list)
    # 添加資料到 Excel
    # e.add_item(stock_prices)
